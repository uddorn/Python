import csv
import openpyxl
from datetime import datetime

def calculate_age(birthdate_str):
    birthdate = datetime.strptime(birthdate_str, "%Y-%m-%d").date()
    today = datetime.today().date()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
    return age

def main():
    csv_filename = "employees.csv"
    xlsx_filename = "employees.xlsx"
    
    employees = []

    try:
        with open(csv_filename, mode='r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                employees.append(row)
    except FileNotFoundError:
        print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV.")
        return
    except Exception as e:
        print("Повідомлення про відсутність, або проблеми при відкритті файлу CSV.")
        return

    try:
        wb = openpyxl.Workbook()

        wb.remove(wb.active)

        sheet_all = wb.create_sheet("all")
        sheet_younger_18 = wb.create_sheet("younger_18")
        sheet_18_45 = wb.create_sheet("18-45")
        sheet_45_70 = wb.create_sheet("45-70")
        sheet_older_70 = wb.create_sheet("older_70")

        headers_all = list(employees[0].keys())
        sheet_all.append(headers_all)

        headers_age_sheets = ["№", "Прізвище", "Ім’я", "По батькові", "Дата народження", "Вік"]
        for sheet in [sheet_younger_18, sheet_18_45, sheet_45_70, sheet_older_70]:
            sheet.append(headers_age_sheets)

        counters = {
            "younger_18": 1,
            "18-45": 1,
            "45-70": 1,
            "older_70": 1
        }

        for row in employees:
            sheet_all.append(list(row.values()))

            age = calculate_age(row["Дата народження"])
            short_row = [
                0, 
                row["Прізвище"],
                row["Ім’я"],
                row["По батькові"],
                row["Дата народження"],
                age
            ]

            if age < 18:
                short_row[0] = counters["younger_18"]
                sheet_younger_18.append(short_row)
                counters["younger_18"] += 1
            elif 18 <= age <= 45:
                short_row[0] = counters["18-45"]
                sheet_18_45.append(short_row)
                counters["18-45"] += 1
            elif 45 < age <= 70:
                short_row[0] = counters["45-70"]
                sheet_45_70.append(short_row)
                counters["45-70"] += 1
            elif age > 70:
                short_row[0] = counters["older_70"]
                sheet_older_70.append(short_row)
                counters["older_70"] += 1

        wb.save(xlsx_filename)
        print("Ok")
        
    except Exception as e:
        print("Повідомлення про неможливість створення XLSX файлу.")

if __name__ == "__main__":
    main()